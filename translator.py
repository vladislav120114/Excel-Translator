from deep_translator import GoogleTranslator
import openpyxl
import os


print("Введите путь до папки с файлами для перевода:")
path = input()
dir = rf"{path}"
print("В папке с файлами появится папка 'Перевод' в которой будут сохранены переведенные файлы")
files = os.listdir(dir)

translator = GoogleTranslator(source='auto', target='ru')

def main(files):
    for file in files[::-1]:
        glos = {}
        try:
            name = file.split(') ')[1]
        except:
            continue
        date = f'{file.split(') ')[0]})'
        glos[name] = translator.translate(name)
        fname = glos[name]
        if os.path.exists(f"{dir}\\Перевод\\{date} {fname}"):
            print(f"{fname} уже переведен")
            continue
        elif file.endswith(".xlsx"):
            print(file)
            wb1 = openpyxl.load_workbook(os.path.join(dir, file))
            for sheet in wb1.sheetnames:
                ws1 = wb1[sheet]
                for row in ws1.iter_rows():
                    for cell in row:
                        if type(cell.value) is str:
                            cell.value = cell.value.replace(',', '.')
                            try: cell.value = float(cell.value)
                            except:
                                if cell.value not in glos:
                                    glos[cell.value] = translator.translate(cell.value)
                                    cell.value = glos[cell.value]
                                else:
                                    cell.value = glos[cell.value]
            if not os.path.exists(f"{dir}\\Перевод"):
                os.makedirs(f"{dir}\\Перевод")
            print(fname)
            wb1.save(f"{dir}\\Перевод\\{date} {fname}")
            print("---")

main(files)